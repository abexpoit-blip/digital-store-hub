#!/usr/bin/env python3
"""
Delivery Format Choice Patch (V1)
---------------------------------
User-কে delivery-র আগে choice দেয়: 📊 Excel (.xlsx) বা 📝 Text (.txt)
- Format একদম identical (UID | PASS | COOKIES 3 column / 3 line)
- delivery_archive-এ save (missing ছিল regular stock-এ)
- openpyxl fail হলে auto TXT fallback

Deploy:
  python3 vps-admin/apply-delivery-format-patch.py
  pm2 restart nexus-bot
"""
import re, shutil, time, os, sys

STORE = "/root/store.py"
BACKUP = f"{STORE}.bak.deliveryfmt.{int(time.time())}"

if not os.path.exists(STORE):
    print(f"❌ {STORE} not found"); sys.exit(1)

src = open(STORE, "r", encoding="utf-8").read()

# --- 1) Skip if already patched ---
MARK = "# [DELIVERY_FORMAT_PATCH_V1]"
if MARK in src:
    print("ℹ️  Already patched. Nothing to do.")
    sys.exit(0)

# --- 2) Backup ---
shutil.copy(STORE, BACKUP)
print(f"✅ Backup: {BACKUP}")

# --- 3) Locate the exact delivery block and replace ---
# The block we replace starts at:
#   _lbl = {"fb61":..}.get(cat, cat.upper()); await m.answer(f"🆔 **আপনার {_lbl} প্রোডাক্ট:**")
# ...and ends after:
#   await show_dashboard_ui(...)
#
# We keep balance cut + sale insert + tempid warn, but REPLACE the raw text
# dump with an inline keyboard, and store items in a module dict for callbacks.

OLD = '''        _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(cat, cat.upper()); await m.answer(f"🆔 **আপনার {_lbl} প্রোডাক্ট:**")
        # [TEMPID_WARN_PATCH]
        if cat == "tempid":
            await m.answer(
                "⚠️ **Temp ID — গুরুত্বপূর্ণ নিয়ম**\\n\\n"
                "⏱ Replace time: **2 ঘণ্টা**\\n\\n"
                "❌ Verify হয়ে গেলে replace **হবে না**\\n"
                "✅ শুধু **login issue** হলে replace সম্ভব\\n\\n"
                "নিয়মের বাইরে replace request দিলে **reject** করা হবে।"
            )

        buffer = ""
        for i in items:
            cursor.execute("DELETE FROM stock WHERE id = ?", (i[0],))
            buffer += i[1] + "\\n\\n"
            if len(buffer) > 3000: await m.answer(buffer); buffer = ""
        if buffer: await m.answer(buffer)
        conn.commit()
        conn.close()

        report_time = "2 Hours" if qty < 10 else "6 Hours"
        await m.answer(f"✅ **সফল!**\\n⏱ রিপোর্ট টাইম: {report_time}\\n🔐 লগইন গ্যারান্টি।")'''

NEW = '''        _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(cat, cat.upper())
        # [DELIVERY_FORMAT_PATCH_V1] — ask format before dumping
        # Get sale_id (last inserted), delete stock, archive, then ask format
        _sale_id = cursor.execute("SELECT last_insert_rowid()").fetchone()[0]
        _now_ts = int(__import__("time").time())
        _uname = (f"@{m.from_user.username}" if m.from_user.username else None)
        _delivered = []
        for i in items:
            cursor.execute("DELETE FROM stock WHERE id = ?", (i[0],))
            _delivered.append((i[0], i[1]))
            try:
                conn.execute(
                    "INSERT INTO delivery_archive (sale_id, user_id, username, category, stock_id, data, source, delivered_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (_sale_id, m.from_user.id, _uname, cat, i[0], i[1], 'bot', _now_ts),
                )
            except Exception:
                pass
        conn.commit()
        conn.close()

        # Stash for callback
        try:
            _PENDING_DELIVERY[_sale_id] = {
                "user_id": m.from_user.id,
                "cat": cat, "lbl": _lbl, "qty": qty,
                "items": _delivered,
                "ts": _now_ts,
            }
        except NameError:
            pass

        if cat == "tempid":
            await m.answer(
                "⚠️ **Temp ID — গুরুত্বপূর্ণ নিয়ম**\\n\\n"
                "⏱ Replace time: **2 ঘণ্টা**\\n\\n"
                "❌ Verify হয়ে গেলে replace **হবে না**\\n"
                "✅ শুধু **login issue** হলে replace সম্ভব\\n\\n"
                "নিয়মের বাইরে replace request দিলে **reject** করা হবে।"
            )

        _kb = types.InlineKeyboardMarkup(inline_keyboard=[[
            types.InlineKeyboardButton(text="📊 Excel (.xlsx)", callback_data=f"dfmt:xlsx:{_sale_id}"),
            types.InlineKeyboardButton(text="📝 Text (.txt)",   callback_data=f"dfmt:txt:{_sale_id}"),
        ]])
        report_time = "2 Hours" if qty < 10 else "6 Hours"
        await m.answer(
            f"✅ **পেমেন্ট সফল!** — {_lbl} × {qty}\\n"
            f"⏱ রিপোর্ট টাইম: {report_time} • 🔐 লগইন গ্যারান্টি\\n\\n"
            f"📥 নিচ থেকে delivery format বেছে নিন:",
            reply_markup=_kb,
        )'''

if OLD not in src:
    print("❌ Target delivery block not found. Was store.py already modified?")
    print("   Restore from backup:", BACKUP)
    sys.exit(2)

src = src.replace(OLD, NEW, 1)
print("✅ Delivery block patched (format selector injected)")

# --- 4) Inject module-level pending dict + callback handlers ---
# Find a safe injection point — right after the last `@dp.callback_query` block
# is risky; simpler: append to end of file (aiogram picks up decorated handlers).

APPEND = '''

# ============================================================
# [DELIVERY_FORMAT_PATCH_V1] — Excel/TXT delivery handlers
# ============================================================
_PENDING_DELIVERY = {}

def _fmt_txt(items, lbl, qty):
    lines = [f"=== {lbl} × {qty} ==="]
    for idx, (_sid, raw) in enumerate(items, 1):
        parts = (raw or "").split(" ")
        uid = parts[0] if parts else ""
        pw  = parts[1] if len(parts) > 1 else ""
        ck  = " ".join(parts[2:]) if len(parts) > 2 else ""
        lines.append(f"\\n--- #{idx} ---\\nUID: {uid}\\nPASS: {pw}\\nCOOKIES: {ck}")
    return ("\\n".join(lines)).encode("utf-8")

def _fmt_xlsx(items, lbl, qty):
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = lbl[:31] or "Delivery"
        hdr = ["#", "UID", "PASS", "COOKIES"]
        ws.append(hdr)
        for c in range(1, 5):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for idx, (_sid, raw) in enumerate(items, 1):
            parts = (raw or "").split(" ")
            uid = parts[0] if parts else ""
            pw  = parts[1] if len(parts) > 1 else ""
            ck  = " ".join(parts[2:]) if len(parts) > 2 else ""
            ws.append([idx, uid, pw, ck])
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 90
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.read(), None
    except Exception as e:
        return None, str(e)

@dp.callback_query(F.data.startswith("dfmt:"))
async def _delivery_format_cb(c: types.CallbackQuery):
    try:
        _, fmt, sid_s = c.data.split(":", 2)
        sid = int(sid_s)
    except Exception:
        return await c.answer("Invalid", show_alert=True)

    meta = _PENDING_DELIVERY.get(sid)
    if not meta:
        # Fallback: pull from archive if bot restarted
        try:
            _cn = sqlite3.connect('/root/store.db')
            _rows = _cn.execute(
                "SELECT stock_id, data, category FROM delivery_archive WHERE sale_id=? ORDER BY id ASC",
                (sid,)
            ).fetchall()
            _cn.close()
            if not _rows: return await c.answer("⚠️ Expired. Admin কে জানান।", show_alert=True)
            _cat = _rows[0][2] or "ITEM"
            _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID","ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}.get(_cat, _cat.upper())
            meta = {"user_id": c.from_user.id, "cat": _cat, "lbl": _lbl,
                    "qty": len(_rows), "items": [(r[0], r[1]) for r in _rows]}
        except Exception:
            return await c.answer("⚠️ Data not found", show_alert=True)

    if meta.get("user_id") and meta["user_id"] != c.from_user.id:
        return await c.answer("⛔ এটা আপনার order না", show_alert=True)

    await c.answer(f"⏳ {fmt.upper()} generate হচ্ছে...")

    lbl = meta["lbl"]; qty = meta["qty"]; items = meta["items"]
    fname_base = f"order-{sid}-{meta['cat']}"

    if fmt == "xlsx":
        data, err = _fmt_xlsx(items, lbl, qty)
        if data is None:
            # fallback to txt
            data = _fmt_txt(items, lbl, qty)
            fname = f"{fname_base}.txt"
            await c.message.answer(f"⚠️ Excel generate fail ({err}) — TXT পাঠানো হলো।")
        else:
            fname = f"{fname_base}.xlsx"
    else:
        data = _fmt_txt(items, lbl, qty)
        fname = f"{fname_base}.txt"

    try:
        from aiogram.types import BufferedInputFile
        await c.message.answer_document(
            BufferedInputFile(data, filename=fname),
            caption=f"📦 {lbl} × {qty}\\n🆔 Order #{sid}\\n\\n💾 এই file বট চ্যাটে permanent — যখন খুশি re-download করতে পারবেন।"
        )
        # Remove keyboard from previous message
        try: await c.message.edit_reply_markup(reply_markup=None)
        except Exception: pass
        # Cleanup memory after successful send
        _PENDING_DELIVERY.pop(sid, None)
    except Exception as e:
        await c.message.answer(f"❌ File পাঠাতে সমস্যা: {e}")
'''

src += APPEND
print("✅ Callback handlers appended")

# --- 5) Write ---
open(STORE, "w", encoding="utf-8").write(src)

# --- 6) Syntax check ---
import py_compile
try:
    py_compile.compile(STORE, doraise=True)
    print("✅ Syntax OK")
except py_compile.PyCompileError as e:
    print("❌ Syntax error! Restoring backup...")
    shutil.copy(BACKUP, STORE)
    print(e); sys.exit(3)

print("\n🎉 Done!")
print("Deploy: pm2 restart nexus-bot && pm2 logs nexus-bot --lines 30 --nostream")
