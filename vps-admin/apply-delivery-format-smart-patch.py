#!/usr/bin/env python3
"""
Delivery Format Smart Hybrid Patch (V2)
----------------------------------------
V1-এর উপরে optimization:

✅ Async file generation — openpyxl asyncio.to_thread()-এ চলে, bot event loop
   block হয় না (100+ concurrent user handle করতে পারে)
✅ Zero VPS storage — file BytesIO-তে RAM-এ বানায়, Telegram-এ পাঠিয়ে RAM free
✅ Telegram CDN permanent hosting — user forever access, VPS zero cost
✅ Memory guard — _PENDING_DELIVERY-এ TTL cleanup (max 200 items, 1hr expire)
✅ Debounce — same button 3sec-এ ২বার click করলে ignore
✅ Auto-fallback — archive-এ pull করে (bot restart-এ ও কাজ করে)
✅ Timing log — কোন step slow বুঝতে

Deploy:
  cd /root/digital-store-hub
  git pull
  python3 vps-admin/apply-delivery-format-smart-patch.py
  pm2 restart nexus-bot
"""
import re, shutil, time, os, sys

STORE = "/root/store.py"
BACKUP = f"{STORE}.bak.deliveryfmt2.{int(time.time())}"

if not os.path.exists(STORE):
    print(f"❌ {STORE} not found"); sys.exit(1)

src = open(STORE, "r", encoding="utf-8").read()

MARK_V1 = "# [DELIVERY_FORMAT_PATCH_V1]"
MARK_V2 = "# [DELIVERY_FORMAT_PATCH_V2]"

if MARK_V2 in src:
    print("ℹ️  V2 already applied. Nothing to do.")
    sys.exit(0)

if MARK_V1 not in src:
    print("❌ V1 not found. Run apply-delivery-format-patch.py first.")
    sys.exit(1)

shutil.copy(STORE, BACKUP)
print(f"✅ Backup: {BACKUP}")

# --- Replace V1 handler block with V2 smart version ---
# Find the V1 appended section (from marker to end of file)
v1_start = src.find("# ============================================================\n# [DELIVERY_FORMAT_PATCH_V1] — Excel/TXT delivery handlers")
if v1_start < 0:
    print("❌ V1 handler block not found"); sys.exit(2)

src_before = src[:v1_start].rstrip() + "\n"

V2_BLOCK = '''

# ============================================================
# [DELIVERY_FORMAT_PATCH_V2] — Smart Hybrid Excel/TXT delivery
# Async file gen, zero VPS storage, Telegram CDN hosting,
# memory TTL guard, debounce, auto-fallback.
# ============================================================
import asyncio as _asyncio_dl
import time as _time_dl

_PENDING_DELIVERY = {}      # sale_id -> {user_id, cat, lbl, qty, items, ts}
_LAST_CLICK = {}            # (user_id, sale_id) -> ts (debounce)
_PENDING_MAX = 200          # cap memory
_PENDING_TTL = 3600         # 1 hour

def _pending_gc():
    """Drop expired / oldest entries — keeps VPS RAM flat."""
    try:
        now = _time_dl.time()
        # TTL cleanup
        expired = [k for k, v in _PENDING_DELIVERY.items()
                   if now - v.get("ts", 0) > _PENDING_TTL]
        for k in expired: _PENDING_DELIVERY.pop(k, None)
        # Size cap — drop oldest
        if len(_PENDING_DELIVERY) > _PENDING_MAX:
            oldest = sorted(_PENDING_DELIVERY.items(),
                            key=lambda kv: kv[1].get("ts", 0))
            for k, _ in oldest[:len(_PENDING_DELIVERY) - _PENDING_MAX]:
                _PENDING_DELIVERY.pop(k, None)
    except Exception:
        pass

def _fmt_txt_sync(items, lbl, qty):
    lines = [f"=== {lbl} × {qty} ==="]
    for idx, (_sid, raw) in enumerate(items, 1):
        parts = (raw or "").split(" ")
        uid = parts[0] if parts else ""
        pw  = parts[1] if len(parts) > 1 else ""
        ck  = " ".join(parts[2:]) if len(parts) > 2 else ""
        lines.append(f"\\n--- #{idx} ---\\nUID: {uid}\\nPASS: {pw}\\nCOOKIES: {ck}")
    return ("\\n".join(lines)).encode("utf-8")

def _fmt_xlsx_sync(items, lbl, qty):
    """CPU-bound — call via asyncio.to_thread so event loop stays free."""
    import openpyxl, io
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(write_only=False)
    ws = wb.active; ws.title = (lbl[:31] or "Delivery")
    ws.append(["#", "UID", "PASS", "COOKIES"])
    _hdr_font = Font(bold=True, color="FFFFFF")
    _hdr_fill = PatternFill("solid", start_color="1F4E78")
    _center = Alignment(horizontal="center")
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = _hdr_font; cell.fill = _hdr_fill; cell.alignment = _center
    for idx, (_sid, raw) in enumerate(items, 1):
        parts = (raw or "").split(" ")
        uid = parts[0] if parts else ""
        pw  = parts[1] if len(parts) > 1 else ""
        ck  = " ".join(parts[2:]) if len(parts) > 2 else ""
        ws.append([idx, uid, pw, ck])
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 90
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

@dp.callback_query(F.data.startswith("dfmt:"))
async def _delivery_format_cb(c: types.CallbackQuery):
    _t0 = _time_dl.time()
    try:
        _, fmt, sid_s = c.data.split(":", 2)
        sid = int(sid_s)
    except Exception:
        return await c.answer("Invalid", show_alert=True)

    # Debounce — same user + sale within 3s → ignore
    _dk = (c.from_user.id, sid)
    if _time_dl.time() - _LAST_CLICK.get(_dk, 0) < 3:
        return await c.answer("⏳ একটু অপেক্ষা করুন...")
    _LAST_CLICK[_dk] = _time_dl.time()
    if len(_LAST_CLICK) > 500: _LAST_CLICK.clear()

    _pending_gc()
    meta = _PENDING_DELIVERY.get(sid)

    # Fallback: pull from archive (works after bot restart too)
    if not meta:
        try:
            _cn = sqlite3.connect('/root/store.db')
            _rows = _cn.execute(
                "SELECT stock_id, data, category, user_id FROM delivery_archive "
                "WHERE sale_id=? ORDER BY id ASC", (sid,)
            ).fetchall()
            _cn.close()
            if not _rows:
                return await c.answer("⚠️ Data নেই। Admin কে জানান।", show_alert=True)
            _cat = _rows[0][2] or "ITEM"
            _owner = _rows[0][3]
            _lbl = {"fb61":"FB 61","fb1000":"FB 1000","tempid":"Temp ID",
                    "ig":"Instagram","fb":"Facebook","bmig":"BM IG","bmfb":"BM FB"}\
                    .get(_cat, _cat.upper())
            meta = {"user_id": _owner, "cat": _cat, "lbl": _lbl,
                    "qty": len(_rows),
                    "items": [(r[0], r[1]) for r in _rows],
                    "ts": _time_dl.time()}
        except Exception as e:
            return await c.answer(f"⚠️ Load fail", show_alert=True)

    if meta.get("user_id") and meta["user_id"] != c.from_user.id:
        return await c.answer("⛔ এটা আপনার order না", show_alert=True)

    await c.answer(f"⏳ {fmt.upper()} তৈরি হচ্ছে...")

    lbl = meta["lbl"]; qty = meta["qty"]; items = meta["items"]
    fname_base = f"order-{sid}-{meta['cat']}"

    # Generate file OFF the event loop → bot stays responsive
    try:
        if fmt == "xlsx":
            try:
                data = await _asyncio_dl.to_thread(_fmt_xlsx_sync, items, lbl, qty)
                fname = f"{fname_base}.xlsx"
            except Exception as _xe:
                # Fallback to TXT so user never gets stuck
                data = await _asyncio_dl.to_thread(_fmt_txt_sync, items, lbl, qty)
                fname = f"{fname_base}.txt"
                await c.message.answer(f"⚠️ Excel unavailable — TXT পাঠানো হলো।")
        else:
            data = await _asyncio_dl.to_thread(_fmt_txt_sync, items, lbl, qty)
            fname = f"{fname_base}.txt"
    except Exception as e:
        return await c.message.answer(f"❌ Generate fail: {e}")

    _gen_ms = int((_time_dl.time() - _t0) * 1000)

    try:
        from aiogram.types import BufferedInputFile
        await c.message.answer_document(
            BufferedInputFile(data, filename=fname),
            caption=(
                f"📦 {lbl} × {qty}  •  🆔 Order #{sid}\\n"
                f"💾 Telegram-এ permanent stored — যখন খুশি re-download।"
            )
        )
        try: await c.message.edit_reply_markup(reply_markup=None)
        except Exception: pass
        # Free RAM immediately after successful send
        _PENDING_DELIVERY.pop(sid, None)
        del data
        _total_ms = int((_time_dl.time() - _t0) * 1000)
        print(f"[delivery] sale={sid} fmt={fmt} qty={qty} gen={_gen_ms}ms total={_total_ms}ms")
    except Exception as e:
        await c.message.answer(f"❌ পাঠাতে সমস্যা: {e}")
'''

new_src = src_before + V2_BLOCK

# Syntax check via write + compile
open(STORE, "w", encoding="utf-8").write(new_src)

import py_compile
try:
    py_compile.compile(STORE, doraise=True)
    print("✅ Syntax OK")
except py_compile.PyCompileError as e:
    print("❌ Syntax error! Restoring backup...")
    shutil.copy(BACKUP, STORE)
    print(e); sys.exit(3)

print("\n🎉 V2 Smart Hybrid applied!")
print("• Async file gen (bot never blocks)")
print("• Zero VPS storage (Telegram CDN hosts forever)")
print("• Memory TTL guard (RAM stays flat)")
print("• Debounce + auto-fallback")
print("\nDeploy: pm2 restart nexus-bot && pm2 logs nexus-bot --lines 30 --nostream")
